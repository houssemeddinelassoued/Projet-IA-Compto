import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Chemin du fichier CSV
csv_file = r'q:\National Quantum\2-PROJETS\Compto-cie\projet-ia-compto\files\Excel Comment créer un rapprochement bancaire(Suivi enregistrement Banque).csv'

# Chemin du fichier Excel à créer
excel_file = csv_file.replace('.csv', '.xlsx')

# Lire le fichier CSV en ligne par ligne
with open(csv_file, 'r', encoding='utf-8') as f:
    lines = f.readlines()

# Créer un workbook
wb = Workbook()
ws = wb.active
ws.title = 'Rapprochement Bancaire'

# Écrire les lignes du CSV dans Excel
for row_idx, line in enumerate(lines, 1):
    values = line.rstrip('\n').split(';')
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value.strip() if value.strip() else None)
        
        # Formatage des en-têtes
        if row_idx in [1, 5]:  # Lignes d'en-tête
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Formatage des nombres
        if row_idx > 5 and col_idx in [5, 6, 7]:  # Colonnes de montants
            try:
                cell.number_format = '#,##0.00'
            except:
                pass

# Ajuster les largeurs de colonnes
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 12

# Sauvegarder le fichier
wb.save(excel_file)

print(f'✓ Fichier Excel créé avec succès')
print(f'  Fichier: {excel_file}')
print(f'  Format: Rapprochement bancaire Compto - Juin 2024')
